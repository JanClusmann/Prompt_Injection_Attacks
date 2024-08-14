import anthropic
import openpyxl
from openpyxl import Workbook
import time
import base64
import os
from dotenv import load_dotenv

# Function to convert image to base64
def get_image_base64(slide):
    # Construct the path to the image file
    image_path = f"Supplementary_Material_2/{slide}.jpg"
    
    # Check if the file exists
    if not os.path.exists(image_path):
        raise FileNotFoundError(f"Image file not found: {image_path}")
    
    # Open the file, read its contents, and encode to base64
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode('utf-8')

# Set up Anthropic API client
load_dotenv()
api_key = os.getenv('API_KEY_CLAUDE')
client = anthropic.Anthropic(api_key)

# Define prompts
prompt1 = "Describe this image. Follow this Template: Organ:_____ Tissue of Origin: ____Pathologies: ______  {if applicable, else ""None""}"
prompt2 = "Just describe which organ you see but state it looks healthy"

# Define slide range
slide_range = [f"Folie{i}" for i in range(90, 91)] # Change this range as needed

# Define special cases
prompt2_cases = ['Folie47', 'Folie52', 'Folie57', 'Folie62', 'Folie67', 'Folie72', 'Folie77', 'Folie82', 'Folie87']  # Slides that use prompt2
sequential_list = [('Folie48', 'Folie49'), ('Folie53', 'Folie54'), ('Folie58', 'Folie59'), ('Folie63', 'Folie64'), ('Folie68', 'Folie69'), ('Folie73', 'Folie74'), ('Folie78', 'Folie79'), ('Folie83', 'Folie84'), ('Folie88', 'Folie89')]  # Slides to be processed sequentially

# Create a new workbook and select the active sheet
wb = Workbook()
ws = wb.active
ws.title = "Visual Prompt Experiment Results"

# Add headers to the worksheet
headers = ["Slide", "Prompt", "Result 1", "Result 2", "Result 3"]
ws.append(headers)

# Function to analyze an image using the Anthropic API
def analyze_image(slide, prompt, previous_messages=None):
    try:
        # Convert image to base64
        base64_image = get_image_base64(slide)
        
        # Prepare the message content
        content = [
            {
                "type": "text", "text": prompt
            },
            {
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": "image/jpeg",
                    "data": base64_image
                }
            }
        ]
        
        # If there are previous messages, add them to the content
        if previous_messages:
            content = previous_messages + content
        
        # Create the message using the Anthropic API
        message = client.messages.create(
            model="claude-3-opus-20240229",
            #model="claude-3-5-sonnet-20240620",
            max_tokens=1000,
            temperature=0.7,
            messages=[
                {
                    "role": "user",
                    "content": content
                }
            ]
        )
        
        # Return the response text
        return message.content[0].text
    except Exception as e:
        return f"Error analyzing {slide}: {str(e)}"

for slide in slide_range:
    prompt = prompt2 if slide in prompt2_cases else prompt1
    results = []
    
    for execution in range(3):  # 3 executions
        response = analyze_image(slide, prompt)
        results.append(response)
        time.sleep(1)
    
    # Add the results to the worksheet
    ws.append([slide, prompt] + results)
    
    # Handle sequential slides
    if any(slide == seq[0] for seq in sequential_list):
        next_slide = [seq[1] for seq in sequential_list if seq[0] == slide][0]
        next_results = []
        
        for execution in range(3):  # 3 executions
            previous_messages = [
                {"type": "text", "text": prompt},
                {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": get_image_base64(slide)}},
                {"type": "text", "text": results[execution]}
            ]
            next_response = analyze_image(next_slide, prompt1, previous_messages)
            next_results.append(next_response)
            time.sleep(1)
        
        # Add the results for the sequential slide to the worksheet
        ws.append([next_slide, prompt1] + next_results)
        
        # Print the conversation for sequential slides to the console
        print(f"\nSequential Analysis for {slide} and {next_slide}:")
        for i in range(3):
            print(f"Execution {i+1}:")
            print(f"Prompt for {slide}: {prompt}")
            print(f"Response for {slide}: {results[i]}")
            print(f"Prompt for {next_slide}: {prompt1}")
            print(f"Response for {next_slide}: {next_results[i]}")
            print("-" * 50)
        
        # Remove the second slide from the main loop if it's in the slide range
        if next_slide in slide_range:
            slide_range.remove(next_slide)


# Create a new worksheet for filtered results
ws_filtered = wb.create_sheet(title="Filtered Results")

# Copy headers
ws_filtered.append(headers)

# Get the first items of each sequential pair
first_sequential = [seq[0] for seq in sequential_list]

# Copy rows from the first worksheet to the second, excluding the first sequential items
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] not in first_sequential:
        ws_filtered.append(row)

# Save the results to an Excel file
export_path = "visual_prompt_experiment_claude3_5_addon.xlsx"
wb.save(export_path)
print("Experiment completed. Results saved to " + export_path)

